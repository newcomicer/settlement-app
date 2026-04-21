import os
import io
import math
import traceback
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def parse_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return {'error': f'無法讀取 Excel 檔案：{e}'}

    # 找報名資料工作表
    reg_sheet = next((n for n in wb.sheetnames if '報名' in n), None)
    if not reg_sheet:
        return {'error': f'找不到「報名資料」工作表（現有工作表：{", ".join(wb.sheetnames)}）'}

    ws = wb[reg_sheet]
    headers = [c.value for c in ws[1]]

    def colidx(keyword):
        return next((i for i, h in enumerate(headers) if h and keyword in str(h)), None)

    def safe_num(val):
        if val is None:
            return 0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0

    # 讀取所有資料列（有日期的才是真實資料列）
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not all_rows:
        return {'error': 'Excel 沒有資料列'}

    data_rows = [r for r in all_rows if r[0] is not None]
    if not data_rows:
        return {'error': '找不到有效的報名資料列（第一欄需有日期）'}

    # ── 合計列偵測（最後一個無日期但有實繳金額的列）──
    paid_col = colidx('實繳金額')
    totals_row = None
    if paid_col is not None:
        for row in reversed(all_rows):
            if row[0] is None and safe_num(row[paid_col]) != 0:
                totals_row = row
                break

    # ── 財務數字：優先用合計列，否則逐訂單加總 ──
    order_col = 1  # 訂單編號通常在第2欄
    # 郵寄費：可能有多欄（本島、離島、現場報到等），全部加總
    # 支援欄位名：「郵寄費用...金額」、「郵寄報到...金額」、「現場報到...金額」
    postal_cols  = [i for i, h in enumerate(headers)
                    if h and ('郵寄' in str(h) or '現場報到' in str(h)) and '金額' in str(h)]
    postal_col   = postal_cols[0] if len(postal_cols) == 1 else None  # 單欄時沿用舊邏輯
    chip_col     = colidx('晶片押金訂單總金額')
    refund_col   = colidx('退費手續費總金額')
    downgrade_col = colidx('降組不退費總金額')
    atm_col      = colidx('ATM虛擬帳號手續費')
    cvs_col      = colidx('超商繳款手續費')
    atm_new_col  = colidx('藍新ATM')
    cvs_new_col  = colidx('藍新超商')

    if totals_row:
        actual_paid  = safe_num(totals_row[paid_col])
        postal       = sum(safe_num(totals_row[c]) for c in postal_cols)
        chip_deposit = safe_num(totals_row[chip_col])     if chip_col      else 0
        refund_fee   = safe_num(totals_row[refund_col])   if refund_col    else 0
        downgrade    = safe_num(totals_row[downgrade_col]) if downgrade_col else 0
        atm_fee      = safe_num(totals_row[atm_new_col] if atm_new_col else 0) + \
                       safe_num(totals_row[atm_col]     if atm_col     else 0)
        cvs_fee      = safe_num(totals_row[cvs_new_col] if cvs_new_col else 0) + \
                       safe_num(totals_row[cvs_col]     if cvs_col     else 0)
    else:
        # 逐訂單加總（每筆訂單只取第一列，避免重複）
        seen = set()
        actual_paid = postal = chip_deposit = refund_fee = downgrade = atm_fee = cvs_fee = 0
        for row in data_rows:
            oid = row[order_col]
            if oid and oid not in seen:
                seen.add(oid)
                if paid_col:      actual_paid  += safe_num(row[paid_col])
                for c in postal_cols: postal   += safe_num(row[c])
                if chip_col:      chip_deposit += safe_num(row[chip_col])
                if refund_col:    refund_fee   += safe_num(row[refund_col])
                if downgrade_col: downgrade    += safe_num(row[downgrade_col])
                if atm_col:       atm_fee      += safe_num(row[atm_col])
                if cvs_col:       cvs_fee      += safe_num(row[cvs_col])
                if atm_new_col:   atm_fee      += safe_num(row[atm_new_col])
                if cvs_new_col:   cvs_fee      += safe_num(row[cvs_new_col])

    # ── 組別人數：動態偵測 ──
    event_col = colidx('參與項目')
    type_col  = colidx('訂單類型')
    if event_col is None:
        return {'error': '找不到「參與項目」欄位，請確認是否為 iRunner 匯出格式'}

    registration = {}
    for row in data_rows:
        event = row[event_col]
        if event:
            registration[str(event)] = registration.get(str(event), 0) + 1

    # ── 公關人數：優先從「免費名單」工作表，其次從訂單類型判斷 ──
    pr_counts = {k: 0 for k in registration}
    free_keywords = ['免費', '公關', '贊助', 'VIP免費']

    if '免費名單' in wb.sheetnames:
        ws_free = wb['免費名單']
        fh = [c.value for c in ws_free[1]]
        fe_col = next((i for i, h in enumerate(fh) if h and '參與項目' in str(h)), None)
        if fe_col is not None:
            for row in ws_free.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    e = str(row[fe_col]) if row[fe_col] else None
                    if e and e in pr_counts:
                        pr_counts[e] += 1
    elif type_col is not None:
        for row in data_rows:
            otype = str(row[type_col]) if row[type_col] else ''
            event = str(row[event_col]) if row[event_col] else ''
            if event in pr_counts and any(k in otype for k in free_keywords):
                pr_counts[event] += 1

    # ── 加購數量：動態偵測，直接加總所有列（加購是每人獨立的）──
    addon_qty_cols = [
        (i, h) for i, h in enumerate(headers)
        if h and ('加購' in str(h) or '加價購' in str(h)) and '總數量' in str(h)
    ]
    # 同時找對應的金額欄，用來推算單價
    addon_amt_cols = {
        str(h).replace('總數量','').replace('訂單','').strip().rstrip('-').strip():
        next((i2 for i2,h2 in enumerate(headers)
              if h2 and str(h2).replace('總金額','').replace('訂單','').strip().rstrip('-').strip()
              == str(h).replace('總數量','').replace('訂單','').strip().rstrip('-').strip()
              and '金額' in str(h2)), None)
        for i, h in addon_qty_cols
    }
    addons = {}
    addon_pr_auto = {}
    addon_prices_auto = {}
    from collections import Counter
    for i, h in addon_qty_cols:
        name = str(h).replace('總數量', '').replace('訂單', '').strip().rstrip('-').strip()
        total_qty = sum(safe_num(row[i]) for row in data_rows)
        addons[name] = int(total_qty)

        # 公關加購數量：訂單類型含免費關鍵字的列加總
        if type_col is not None:
            pr_qty = sum(
                safe_num(row[i]) for row in data_rows
                if any(kw in str(row[type_col] or '') for kw in free_keywords)
            )
            if pr_qty > 0:
                addon_pr_auto[name] = int(pr_qty)

        # 推算單價：只取有付費（金額>0）且有數量的列，取最常見的單價
        amt_col = addon_amt_cols.get(name)
        if amt_col is not None:
            unit_prices = []
            for row in data_rows:
                qty = safe_num(row[i])
                amt = safe_num(row[amt_col])
                if qty > 0 and amt > 0:
                    unit_prices.append(int(round(amt / qty)))
            if unit_prices:
                addon_prices_auto[name] = Counter(unit_prices).most_common(1)[0][0]

    # ── 組別單價：從報名項目費用欄自動推算（取付費者最常見的金額）──
    reg_prices_auto = {}
    fee_col = colidx('報名項目費用')
    free_keywords = ['免費', '公關', '贊助', 'VIP免費']
    if fee_col is not None and type_col is not None:
        from collections import Counter
        for k in registration:
            prices_seen = []
            for row in data_rows:
                if str(row[event_col] or '') != k:
                    continue
                otype = str(row[type_col] or '')
                if any(kw in otype for kw in free_keywords):
                    continue  # 排除免費名額
                fee = safe_num(row[fee_col])
                if fee > 0:
                    prices_seen.append(int(fee))
            if prices_seen:
                reg_prices_auto[k] = Counter(prices_seen).most_common(1)[0][0]

    return {
        'registration': registration,
        'pr_counts': pr_counts,
        'addons': addons,
        'reg_prices_auto': reg_prices_auto,
        'addon_prices_auto': addon_prices_auto,
        'addon_pr_auto': addon_pr_auto,
        'financials': {
            'actual_paid':  int(actual_paid),
            'postal':       int(postal),
            'chip_deposit': int(chip_deposit),
            'refund_fee':   int(refund_fee),
            'downgrade':    int(downgrade),
            'atm_fee':      int(atm_fee),
            'cvs_fee':      int(cvs_fee),
        },
        'has_totals_row': totals_row is not None,
    }


def calculate_settlement(data):
    reg       = data['registration']    # {group: count}
    pr        = data['pr_overrides']    # {group: pr_count}
    addons    = data['addons']          # {name: qty}
    addon_pr  = data.get('addon_pr', {})
    fin       = data['financials']
    manual    = data['manual']

    reg_prices   = manual.get('reg_prices', {})
    addon_prices = manual.get('addon_prices', {})

    # 報名費
    reg_fee = sum(
        (reg.get(k, 0) - pr.get(k, 0)) * reg_prices.get(k, 0)
        for k in reg
    )

    # 加價購費
    addon_fee = sum(
        (addons.get(k, 0) - addon_pr.get(k, 0)) * addon_prices.get(k, 0)
        for k in addons
    )

    # 計時服務費（全部由使用者自訂）
    total_participants = sum(reg.values())
    billing_count = max(total_participants, math.ceil(total_participants / 1000) * 1000)

    timing_breakdown = []
    timing_subtotal = 0
    for item in manual.get('timing_items', []):
        qty   = float(item.get('qty', 0))
        unit  = float(item.get('unit', 0))
        is_pct = item.get('is_percent', False)
        if is_pct:
            sub = round(timing_subtotal * unit / 100)
        else:
            sub = round(qty * unit)
        timing_breakdown.append({
            'item':       item.get('name', ''),
            'qty':        qty,
            'unit':       unit,
            'subtotal':   sub,
            'is_percent': is_pct,
        })
        timing_subtotal += sub

    timing_total = round(timing_subtotal * 1.05)

    # 方式一
    atm_fee  = -manual.get('atm_fee', 0)
    cvs_fee  = -manual.get('cvs_fee', 0)
    refund   = -(fin.get('refund_fee', 0) + fin.get('downgrade', 0) + manual.get('refund_extra', 0))
    chip     = -fin.get('chip_deposit', 0)
    method1_total = fin['actual_paid'] + atm_fee + cvs_fee + refund + chip - timing_total

    # 方式二
    method2_total = reg_fee + addon_fee + fin.get('postal', 0) + manual.get('overpaid', 0) - timing_total

    # 組別明細
    reg_breakdown = [
        {
            'name':  k,
            'count': reg.get(k, 0),
            'pr':    pr.get(k, 0),
            'price': reg_prices.get(k, 0),
            'total': (reg.get(k, 0) - pr.get(k, 0)) * reg_prices.get(k, 0),
        }
        for k in reg
    ]

    # 加購明細
    addon_breakdown = [
        {
            'name':  k,
            'qty':   addons.get(k, 0),
            'pr':    addon_pr.get(k, 0),
            'price': addon_prices.get(k, 0),
            'total': (addons.get(k, 0) - addon_pr.get(k, 0)) * addon_prices.get(k, 0),
        }
        for k in addons
    ]

    return {
        'method1': {
            'actual_paid':  fin['actual_paid'],
            'atm_fee':      atm_fee,
            'cvs_fee':      cvs_fee,
            'refund':       refund,
            'chip_deposit': chip,
            'total':        method1_total,
        },
        'method2': {
            'reg_fee':   reg_fee,
            'addon_fee': addon_fee,
            'postal':    fin.get('postal', 0),
            'overpaid':  manual.get('overpaid', 0),
            'total':     method2_total,
        },
        'timing_breakdown': timing_breakdown,
        'timing_subtotal':  timing_subtotal,
        'timing_total':     timing_total,
        'reg_breakdown':    reg_breakdown,
        'addon_breakdown':  addon_breakdown,
        'total_participants': total_participants,
        'billing_count':    billing_count,
    }


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/parse-excel', methods=['POST'])
def parse_excel_route():
    if 'file' not in request.files:
        return jsonify({'error': '未上傳檔案'}), 400
    f = request.files['file']
    path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
    f.save(path)
    try:
        result = parse_excel(path)
    except Exception as e:
        return jsonify({'error': f'解析失敗：{e}', 'detail': traceback.format_exc()}), 500
    return jsonify(result)


@app.route('/generate-pdf', methods=['POST'])
def generate_pdf_route():
    try:
        from playwright.sync_api import sync_playwright
        import tempfile

        data = request.json
        settlement = calculate_settlement(data)
        html_str = render_template('pdf_pages.html', data=data, settlement=settlement, manual=data['manual'])

        # 寫入暫存 HTML，讓 Playwright 讀取
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8') as f:
            f.write(html_str)
            tmp_html = f.name

        tmp_pdf = tmp_html.replace('.html', '.pdf')

        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.goto(f'file://{tmp_html}')
            page.wait_for_load_state('networkidle')
            page.pdf(
                path=tmp_pdf,
                format='A4',
                margin={'top': '15mm', 'bottom': '15mm', 'left': '15mm', 'right': '15mm'},
                print_background=True,
            )
            browser.close()

        os.unlink(tmp_html)

        with open(tmp_pdf, 'rb') as f:
            out = io.BytesIO(f.read())
        os.unlink(tmp_pdf)
        out.seek(0)

        name = data.get('manual', {}).get('event_name', '活動')
        return send_file(out, as_attachment=True,
                         download_name=f'費用申請_{name}.pdf',
                         mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': str(e), 'detail': traceback.format_exc()}), 500


if __name__ == '__main__':
    print('✅ 經費結算系統啟動中...')
    print('請開啟瀏覽器前往: http://127.0.0.1:5001')
    app.run(debug=True, host='127.0.0.1', port=5001)
